from pathlib import Path
from colorama import Fore
from nornir.core import Nornir
from nornir.core.exceptions import NornirSubTaskError, NornirExecutionError
from nornir.core.task import Task, AggregatedResult
from nornir_napalm.plugins.tasks import napalm_get
from nornir_netmiko import netmiko_send_command
from openpyxl import Workbook

from modules.utility.excel_exporter import ExcelExporter
from modules.utility.network_info_collector import NetworkInfoCollector
from modules.utility.network_info_parser import NetworkInfoParser
from modules.utility.text_file_exporter import FileExporter


class NetworkInfoExporter:
    """
    Třída která umožňuje exportovat data ze síťových zařízení.

    Args:
        info_collector (NetworkInfoCollector): objekt, který slouží k obdržení komplexních dat ze síťových prvků - např. při paralelním slučování
                                               několika MultiResult objektů z více Nornir podúloh (subtasků) s jiným typem přístupu k NAPALM knihovně.


    Attributes:
         info_collector (NetworkInfoCollector): objekt, který slouží k obdržení komplexních dat ze síťových prvků - např. při paralelním slučování
                                               několika MultiResult objektů z více Nornir podúloh (subtasků) s jiným typem přístupu k NAPALM knihovn

    """

    def __init__(self, info_collector: NetworkInfoCollector):
        self._info_collector = info_collector

    def export_device_facts(self, nornir_devices: Nornir) -> None:
        """
        Export základních údajů jednotlivých zařízení do .xlsx souboru.

        Args:
            nornir_devices (Nornir): filtrovaný Nornir objekt umožňující na daných zařízeních volat nornir úkoly.

        Returns:
            None

        """
        sorted_headers_export = ["hostname", "FQDN", "vendor", "model", "serial_number", "os_version", "uptime", "connection"]
        wider_header_columns = ["os_version", "FQDN"]
        dest_file_path = Path(Path.cwd() / 'export' / "excel" / "facts.xlsx")
        try:
            all_results_aggregation: AggregatedResult = nornir_devices.run(
                task=self._info_collector.get_conn_state_and_device_facts)
            exporter = ExcelExporter(Workbook(), "Základní informace o zařízeních", dest_file_path)
            parser = NetworkInfoParser()
            parsed_data = parser.get_parsed_facts_data(all_results_aggregation)
            exporter.write_header(sorted_headers_export, wider_header_columns)
            exporter.write_data(sorted_headers_export, parsed_data)
            exporter.save_xlsx_file()
        except (NornirExecutionError, OSError, ValueError) as err:
            print(f"{Fore.RED}Export device facts to {dest_file_path.name} failed.")
            print(err)


    def export_interfaces_packet_counters(self, nornir_devices: Nornir) -> None:
        """
         Export statistik týkající se přijatých a vysílaných paketů pro jednotlivá rozhraní síťových zařízení. Statistiky jsou exportovány do .xslx souboru.
         Tx_broadcast a tx_multicast nejsou exportovány vůbec (NAPALM momentálně nepodporuje tyto statistiky - \n
         U vývojáři nedefinovaných regexů vychází vždy daná statistika -1 (platí pro tx_broadcast, tx_multicast a několik paket counterů u Juniper Olive).

         Args:
             nornir_devices (Nornir): filtrovaný Nornir objekt umožňující na daných zařízeních volat nornir úkoly.

         Returns:
             None

         """
        sorted_headers_export = ["interface", "rx_broadcast", "rx_discards", "rx_errors",
                                 "rx_multicast", "rx_octets", "rx_unicast",
                                 "tx_discards", "tx_errors", "tx_octets", "tx_unicast"]
        wider_header_columns = ["interface"]
        dest_file_path = Path(Path.cwd() / 'export' / "excel" / f"packets_counter.xlsx")
        try:
            all_results_aggregation: AggregatedResult = nornir_devices.run(
                task=napalm_get, name="Get interfaces packet counters", getters=["interfaces_counters"])
            parser = NetworkInfoParser()
            parser.parse_interfaces_packet_counters_data(all_results_aggregation)
            first_host = list(all_results_aggregation.keys())[0]
            exporter = ExcelExporter(Workbook(), first_host, dest_file_path)
            for host in all_results_aggregation:
                sheet_name = host
                exporter.create_sheet(sheet_name)
                exporter.change_active_sheet(sheet_name)
                row = 2
                column = 1
                exporter.write_header(sorted_headers_export, wider_header_columns, row, column)
                interfaces_data = all_results_aggregation[host][0].result['interfaces_counters']
                host_interfaces_lst = []
                for interface in sorted(interfaces_data):
                    interface_dict = interfaces_data[interface]
                    interface_dict['interface'] = interface
                    host_interfaces_lst.append(interface_dict)
                exporter.write_data(sorted_headers_export, host_interfaces_lst, row, column)
            exporter.save_xlsx_file()
        except (ValueError, NornirExecutionError) as err:
            print(f"{Fore.RED}Export interfaces packet counters to {dest_file_path.name} failed.")
            print(err)
        except OSError as err:
            print(f"{Fore.RED}Export interfaces packet counters to {dest_file_path.name} failed.")
            print(f"{Fore.RED}Error: Creating directories for specified path {dest_file_path.parent} failed.")


    def export_device_configuration(self, task: Task) -> None:
        """
        Export současné (running) konfigurace síťových prvků do .conf souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/running_configuration/{konkrétní host}.conf (při spuštění skriptu na GNU/Linux).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Returns:
            None

        """
        result = task.run(task=napalm_get, name="Get configuration", getters=["config"])
        if not result.failed:
            running_configuration = result[0].result["config"]['running'].strip()
            file_path = Path(Path.cwd() / 'export' / "running_configuration" / f"{task.host.name}.conf")
            exporter = FileExporter(file_path, running_configuration)
            exporter.export_to_file()
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")

    def export_ipv4_routes(self, task: Task) -> None:
        """
        Export směrovací IPv4 tabulky jednotlivých síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/ip_routes/{konkrétní host}_ipv4.txt (při spuštění skriptu na GNU/Linux).
        Podporovaní výrobci + typy zařízení: Cisco (L3 switche nebo routery), Juniper (routery).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Raises:
            NornirSubTaskError: Výjimka, která nastane, pokud se objeví chyba v nornir úkolu nebo pokud provádíte
                export na nepodporovaných zařízeních.

        Returns:
            None

        """
        command = ""
        if task.host['vendor'] == "cisco" and (
                task.host['dev_type'] == "router" or task.host['dev_type'] == "L3_switch"):
            command = "show ip route"
        elif task.host['vendor'] == "juniper" and task.host['dev_type'] == "router":
            command = "show route"
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} - not implemented for that vendor.")
            raise NornirSubTaskError(f"Function was not implemented for vendor {task.host.name}.", task)
        result = task.run(task=netmiko_send_command, name="Get IP routes", command_string=command)
        if not result.failed:
            ipv4_routes = result[0].result
            if task.host['vendor'] == "juniper":
                parser = NetworkInfoParser()
                ipv4_routes = parser.get_parsed_juniper_routes(result)
            if ipv4_routes != "":
                file_path = Path(Path.cwd() / 'export' / "ip_routes" / f"{task.host.name}_ipv4.txt")
                exporter = FileExporter(file_path, ipv4_routes)
                exporter.export_to_file()
            else:
                print(f"{Fore.RED}{task.host.name}: No IPv4 routes are defined.")
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")

    def export_ipv6_routes(self, task: Task) -> None:
        """
        Export směrovací IPv6 tabulky jednotlivých síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/ip_routes/{konkrétní host}_ipv6.txt (při spuštění skriptu na GNU/Linux).
        Podporovaní výrobci + typy zařízení: Cisco (L3 switche nebo routery), Juniper (routery).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Raises:
            NornirSubTaskError: Výjimka, která nastane, pokud nastana chyba v nornir úkolu nebo pokud provádíte
                export na nepodporovaných zařízeních.

        Returns:
            None

        """
        command = ""
        if task.host['vendor'] == "cisco" and (
                task.host['dev_type'] == "router" or task.host['dev_type'] == "L3_switch"):
            command = "show ipv6 route"
        elif task.host['vendor'] == "juniper" and task.host['dev_type'] == "router":
            command = "show route"
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} - not implemented for that vendor.")
            raise NornirSubTaskError(f"Function was not implemented for vendor {task.host.name}.", task)
        result = task.run(task=netmiko_send_command, name="Get IP routes", command_string=command)
        if not result.failed:
            ipv6_routes = result[0].result
            if task.host['vendor'] == "juniper":
                parser = NetworkInfoParser()
                ipv6_routes = parser.get_parsed_juniper_routes(result, ipv6_routes=True)
            if ipv6_routes != "":
                file_path = Path(Path.cwd() / 'export' / "ip_routes" / f"{task.host.name}_ipv6.txt")
                exporter = FileExporter(file_path, ipv6_routes)
                exporter.export_to_file()
            else:
                print(f"{Fore.RED}{task.host.name}: No IPv6 routes are defined.")
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")

    def export_packet_filter_info(self, task: Task) -> None:
        """
        Export definovaných paketových filtrů (IPv4 i IPv6) u jednotlivých síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/packet_filter/{konkrétní host}.txt (při spuštění skriptu na GNU/Linux).
        Podporovaní výrobci + typy zařízení: Cisco (L3 switche nebo routery), Juniper (routery).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Raises:
            NornirSubTaskError: Výjimka, která nastane, pokud nastana chyba v nornir úkolu nebo pokud provádíte
                export na nepodporovaných zařízeních.

        Returns:
            None

        """
        command = ""
        if task.host['vendor'] == "cisco" and (
                task.host['dev_type'] == "router" or task.host['dev_type'] == "L3_switch"):
            command = "show access-lists"
        elif task.host['vendor'] == "juniper" and task.host['dev_type'] == "router":
            command = "show configuration firewall"
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} - not implemented for that vendor.")
            raise NornirSubTaskError(f"Function was not implemented for vendor {task.host.name}.", task)
        result = task.run(task=netmiko_send_command, name="Get packet filters info", command_string=command)
        if not result.failed:
            packet_filter_info = result[0].result
            if packet_filter_info != "":
                file_path = Path(Path.cwd() / 'export' / "packet_filter" / f"{task.host.name}.txt")
                exporter = FileExporter(file_path, packet_filter_info)
                exporter.export_to_file()
            else:
                print(f"{Fore.RED}{task.host.name}: No packet filter is defined.")
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")
