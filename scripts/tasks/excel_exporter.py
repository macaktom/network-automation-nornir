from pathlib import Path

from nornir.core.task import AggregatedResult, Result, Task
from nornir_napalm.plugins.tasks import napalm_get
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from os import path
import datetime


class ExcelExporter:

    def __init__(self, workbook: Workbook, sheet_title: str, dest_file: Path):
        self._workbook = workbook
        self._dest_file = dest_file
        self._setup_export(sheet_title)

    def _create_parent_folders(self, folder_path: Path) -> None:
        folder_path.mkdir(parents=True, exist_ok=True)

    def _check_file_dest(self) -> None:
        if path.isfile and self._dest_file.suffix == '.xlsx':
            self._create_parent_folders(self._dest_file.parent)
        else:
            raise ValueError("Not valid filepath or extension")

    def _write_data(self, lst_data: list, sorted_headers: list, row_start: int, column_start: int) -> None:
        current_column = column_start
        for host_data in lst_data:
            row_start += 1
            for header in sorted_headers:
                if host_data[header] is None or host_data[header] == "None" or host_data[header] == "":
                    self.ws1.cell(row=row_start, column=current_column).value = "-"
                else:
                    self.ws1.cell(row=row_start, column=current_column).value = host_data[header]
                self.ws1.cell(row=row_start, column=current_column).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
                current_column += 1
            current_column = column_start

    def _write_header(self, headers: list, row_start: int, column_start: int) -> None:
        for header in headers:
            self.ws1.cell(row=row_start, column=column_start).value = header
            self.ws1.cell(row=row_start, column=column_start).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
            if header.lower() == "os_version" or header.lower() == "fqdn":
                self.ws1.column_dimensions[get_column_letter(column_start)].width = 40
            else:
                self.ws1.column_dimensions[get_column_letter(column_start)].width = 15
            column_start += 1

    def _parse_facts_data(self, aggregated_dict_result: AggregatedResult) -> tuple:
        data = []
        sorted_headers_export = ["hostname", "FQDN", "vendor", "serial_number", "os_version", "uptime", "connection"]
        for host in aggregated_dict_result:
            data_dict = aggregated_dict_result[host][1].result['facts']
            data_dict['connection'] = "OK" if aggregated_dict_result[host][2].result else "Failed"
            data_dict['uptime'] = str(datetime.timedelta(seconds=data_dict['uptime']))
            data_dict["FQDN"] = data_dict["fqdn"]
            del data_dict["fqdn"]
            del data_dict["interface_list"]
            version = data_dict['os_version']
            if data_dict['vendor'].lower() == "cisco":
                parsed_version = version.split(",")[1].strip()
                data_dict['os_version'] = parsed_version
            data.append(data_dict)
        print(data)
        return sorted_headers_export, data

    def export_to_xlsx(self, aggregated_dict_result: AggregatedResult, row_start: int = 2, column_start: int = 2) -> None:
        if row_start > 0 and column_start > 0:
            sorted_headers_export, data = self._parse_facts_data(aggregated_dict_result)
            self._write_header(sorted_headers_export, row_start, column_start)
            self._write_data(data, sorted_headers_export, row_start, column_start)
            self._workbook.save(self._dest_file)
        else:
            print("Data were not exported - check specified row_start and column_start (both must be > 0).")

    def _setup_export(self, sheet_title: str) -> None:
        self._check_file_dest()
        self.ws1 = self._workbook.active
        self.ws1.title = sheet_title

    def _get_conn_state_result(self, task: Task) -> Result:
        napalm = task.host.get_connection("napalm", task.nornir.config)
        result = napalm.is_alive()
        return Result(host=task.host, result=result)

    def get_conn_state_and_device_facts(self, task: Task) -> AggregatedResult:
        result = task.run(task=napalm_get, name="Get device basic facts", getters=["facts"])
        result += task.run(task=self._get_conn_state_result, name="Get conn status")
        return AggregatedResult(*result)
