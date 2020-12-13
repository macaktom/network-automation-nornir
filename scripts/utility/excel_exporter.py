from pathlib import Path
from typing import List, Dict

from colorama import Fore
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Třida pro exportování dat do formátu .xlsx.

    Args:
        workbook (Workbook): objekt openpyxl knihovny reprezentující sešit.
        sheet_title (str): název výchozího listu
        dest_file (Path): argument metody, obsahující cestu k exportovanému .xlsx souboru.

    Attributes:
         workbook (Workbook): instanční proměnná, reprezentující sešit v tabulkovém procesoru.
         dest_file (Path): argument metody, obsahující cestu k exportovanému .xlsx souboru.
         sheets (List[Worksheet]): listy tabulkového procesoru, které jsou obsaženy v daném sešitě.
         active_sheet (Worksheet): instanční proměnná reprezentující aktivní list

    """
    def __init__(self, workbook: Workbook, sheet_title: str, dest_file: Path):
        self._workbook = workbook
        self._dest_file = dest_file
        self._sheets = []
        self._setup_export(sheet_title)

    def _create_parent_folders(self, folder_path: Path) -> None:
        """
        Metoda pro vytvoření všech nadřazených složek k definované instanční proměnné dest_file.
        Složky jsou vytvořeny pouze tehdy, pokud nebyly dříve uživatelem vytvořeny.

        Args:
            folder_path (Path): cesta obsahující všechny nadřazené složky, které jsou nutné pro nalezení výsledného .txt souboru

        Raises:
            OSError: Výjimka, která nastane pokud došlo k chybě při vytváření složek z definované cesty folder_path.

        Returns:
            None
        """
        try:
            folder_path.mkdir(parents=True, exist_ok=True)
        except OSError:
            print(f"{Fore.RED}Error: Creating directories for specified path {folder_path} failed.")

    def _check_file_dest(self) -> None:
        """
        Metoda, která zkontroluje, jestli je k instanční proměnné dest_file přiřazena validní cesta k .xlsx souboru.
        Dodatečně vytvoří požadované nadřazené složky, pokud už nejsou vytvořené za pomocí metody _create_parent_folders.

        Raises:
            ValueError: Výjimka, která nastane pokud atribut dest_file není validní cestou k .xlsx souboru.

        Returns:
            None
        """
        if self._dest_file.suffix == '.xlsx':
            self._create_parent_folders(self._dest_file.parent)
        else:
            raise ValueError("Not valid filepath or extension.")

    def write_data(self, sorted_headers: List[str], lst_data: List[Dict[str, str]], row_start: int = 2,
                   column_start: int = 1) -> None:
        """
        Metoda pro zápis dat do aktivního listu.

        Args:
            sorted_headers (List[str]): seřazená hlavička tabulky, určující pořadí zápisu.
            lst_data (List[Dict[str, str]]): data určená k exportu.
            row_start (int): počáteční řádek zápisu. Defaultní hodnota je 2.
            column_start (int): počáteční sloupec zápisu. Defaultní hodnota je 1.

        Returns:
            None
        """
        if row_start > 0 and column_start > 0:
            current_column = column_start
            for data in lst_data:
                row_start += 1
                for header in sorted_headers:
                    if data[header] is None or data[header] == "None" or data[header] == "":
                        self.active_sheet.cell(row=row_start, column=current_column).value = "-"
                    else:
                        self.active_sheet.cell(row=row_start, column=current_column).value = data[header]
                    self.active_sheet.cell(row=row_start, column=current_column).alignment = Alignment(
                        horizontal='center',
                        vertical='center')
                    current_column += 1
                current_column = column_start
        else:
            print(f"{Fore.RED}Error with writing data - check specified row_start and column_start (both must be > 0).")

    def write_header(self, sorted_headers: List[str], wider_header_columns: List[str], row_start: int = 2,
                     column_start: int = 1) -> None:
        """
        Metoda pro zápis hlavičky tabulky do aktivního listu.

        Args:
            sorted_headers (List[str]): seřazená hlavička tabulky
            wider_header_columns (List[str]): list nadpisů, který určuje, jaké sloupce mají být širší.
            row_start (int): počáteční řádek zápisu. Defaultní hodnota je 2.
            column_start (int): počáteční sloupec zápisu. Defaultní hodnota je 1.

        Returns:
            None
        """
        if row_start > 0 and column_start > 0:
            for header in sorted_headers:
                self.active_sheet.cell(row=row_start, column=column_start).value = header
                self.active_sheet.cell(row=row_start, column=column_start).alignment = Alignment(horizontal='center',
                                                                                                 vertical='center')
                if header in wider_header_columns:
                    self.active_sheet.column_dimensions[get_column_letter(column_start)].width = 30
                else:
                    self.active_sheet.column_dimensions[get_column_letter(column_start)].width = 15
                column_start += 1
        else:
            print(f"{Fore.RED}Error with writing header - check specified row_start and column_start (both must be > 0).")

    def save_xlsx_file(self) -> None:
        """
        Metoda pro uložení/vytvoření .xlsx souboru. Lokace souboru je definována atributem dest_file.

        Raises:
            Exception: Vyjímka, která nastane, pokud je problém s uložením/vytvořením výsledného .xlsx souboru.

        Returns:
            None
        """
        try:
            self._workbook.save(self._dest_file)
            print(f"Export data to {self._dest_file.name} was successful.")
        except Exception:
            print(f"{Fore.RED}Export failed. Error while saving .xslx file.")

    def _setup_export(self, sheet_title: str) -> None:
        """
        Metoda pro počáteční nastavení exportu .xlsx souboru (kontrola cesty k souboru, vytvoření defaultního listu, nastavení názvu listu).

        Args:
            sheet_title (str): název výchozího listu

        Returns:
            None
        """
        self._check_file_dest()
        ws1 = self._workbook.active
        self._sheets.append(ws1)
        self.active_sheet = ws1
        ws1.title = sheet_title

    def change_active_sheet(self, sheet_title: str) -> None:
        """
        Metoda, která nastaví daný list jako aktivní.

        Args:
            sheet_title (str): název listu

        Returns:
            None
        """
        if sheet_title:
            for sheet in self._sheets:
                if sheet.title.lower() == sheet_title.lower():
                    self._workbook.active = sheet
                    self.active_sheet = sheet

    def create_sheet(self, sheet_title: str) -> None:
        """
        Metoda pro vytvoření nového listu.

        Args:
            sheet_title (str): název nového listu

        Returns:
            None
        """
        if sheet_title not in self._workbook.sheetnames:
            ws = self._workbook.create_sheet(sheet_title)
            self._sheets.append(ws)
