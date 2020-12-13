import json
from pathlib import Path

from colorama import Fore
from nornir.core import Nornir
from nornir.core.exceptions import NornirSubTaskError
from nornir.core.task import Task, MultiResult, AggregatedResult
from nornir_napalm.plugins.tasks import napalm_get
from nornir_netmiko import netmiko_send_command, netmiko_send_config
from openpyxl import Workbook

from scripts.utility.excel_exporter import ExcelExporter
from scripts.utility.network_info_collector import NetworkInfoCollector
from scripts.utility.network_info_parser import NetworkInfoParser
from scripts.utility.text_file_exporter import TextFileExporter


class NetworkInfoExporter:
    """
    Třída která umožňuje exportovat data ze síťových zařízení.
    """

    def export_device_facts(self, nornir_devices: Nornir, dest_file_path: Path) -> None:
        """
        Export základních údajů jednotlivých zařízení do .xlsx souboru.

        Args:
            nornir_devices (Nornir): filtrovaný Nornir objekt umožňující na daných zařízeních volat nornir úkoly.
            dest_file_path: (Path): cesta k výslednému .xlsx souboru.

        Returns:
            None

        """
        info_collector = NetworkInfoCollector()
        all_results_aggregation: AggregatedResult = nornir_devices.run(
            task=info_collector.get_conn_state_and_device_facts)
        sorted_headers_export = ["hostname", "FQDN", "vendor", "serial_number", "os_version", "uptime", "connection"]
        wider_header_columns = ["os_version", "FQDN"]
        if not all_results_aggregation.failed:
            exporter = ExcelExporter(Workbook(), "Základní informace o zařízeních", dest_file_path)
            parser = NetworkInfoParser()
            parsed_data = parser.get_parsed_facts_data(all_results_aggregation)
            exporter.write_header(sorted_headers_export, wider_header_columns)
            exporter.write_data(sorted_headers_export, parsed_data)
            exporter.save_xlsx_file()
        else:
            print(f"{Fore.RED}Export failed - more in nornir.log")

    def export_interfaces_packet_counters(self, nornir_devices: Nornir, dest_file_path: Path) -> None:
        """
         Export statistik týkající se přijímání a vysílání paketů pro jednotlivá rozhraní síťových zařízení. Statistiky jsou exportovány do .xslx souboru.

         Args:
             nornir_devices (Nornir): filtrovaný Nornir objekt umožňující na daných zařízeních volat nornir úkoly.
             dest_file_path: (Path): cesta k výslednému .xlsx souboru.

         Returns:
             None

         """
        info_collector = NetworkInfoCollector()
        all_results_aggregation: AggregatedResult = nornir_devices.run(
            task=info_collector.get_interfaces_packet_counters)
        sorted_headers_export = ["interface", "rx_broadcast", "rx_discards", "rx_errors",
                                 "rx_multicast", "rx_octets", "rx_unicast", "tx_broadcast",
                                 "tx_discards", "tx_errors", "tx_multicast", "tx_octets", "tx_unicast"]
        wider_header_columns = ["interface"]
        if not all_results_aggregation.failed:
            parser = NetworkInfoParser()
            parser.parse_interfaces_packet_counters_data(all_results_aggregation)
            exporter = ExcelExporter(Workbook(), list(all_results_aggregation.keys())[0], dest_file_path)
            for host in all_results_aggregation:
                sheet_name = host
                exporter.create_sheet(sheet_name)
                exporter.change_active_sheet(sheet_name)
                row = 2
                column = 1
                exporter.write_header(sorted_headers_export, wider_header_columns, row, column)
                intefaces_data = all_results_aggregation[host][1].result['interfaces_counters']
                for interface in sorted(intefaces_data):
                    interface_dict = intefaces_data[interface]
                    interface_dict['interface'] = interface
                    exporter.write_data(sorted_headers_export, [interface_dict], row, column)
                    row += 1
            exporter.save_xlsx_file()
        else:
            print(f"{Fore.RED}Export failed - more in nornir.log")

    def export_device_configuration(self, task: Task) -> None:
        """
        Export současné (running) konfigurace síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/running_configuration/{konkrétní host}.txt (při spuštění skriptu na GNU/Linux).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Returns:
            None

        """
        collector = NetworkInfoCollector()
        result = collector.get_device_configuration(task)
        if not result.failed:
            running_configuration = result[0].result["config"]['running'].strip()
            file_path = Path(Path.cwd() / 'export' / "running_configuration" / f"{task.host.name}.txt")
            exporter = TextFileExporter(file_path, running_configuration)
            exporter.export_to_file()
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")

    def export_ipv4_routes(self, task: Task) -> None:
        """
        Export směrovací IPv4 tabulky jednotlivých síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/ip_routes/{konkrétní host}_ipv4.txt (při spuštění skriptu na GNU/Linux).
        Podporovaní výrobci + typy zařízení: Cisco (L3 switche nebo routery), Juniper (routery).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Raises:
            NornirSubTaskError: Výjimka, která nastane, pokud se objeví chyba v nornir úkolu nebo pokud provádíte
                export na nepodporovaných zařízeních.

        Returns:
            None

        """
        command = ""
        if task.host['vendor'] == "cisco" and (
                task.host['dev_type'] == "router" or task.host['dev_type'] == "L3_switch"):
            command = "show ip route"
        elif task.host['vendor'] == "juniper" and task.host['dev_type'] == "router":
            command = "show route"
        else:
            raise NornirSubTaskError("Function was not implemented for particular vendor.", task)
        result = task.run(task=netmiko_send_command, name="Get IP routes", command_string=command)
        if not result.failed:
            ipv4_routes = result[0].result
            if task.host['vendor'] == "juniper":
                parser = NetworkInfoParser()
                ipv4_routes = parser.get_parsed_juniper_routes(result)
            if ipv4_routes != "":
                file_path = Path(Path.cwd() / 'export' / "ip_routes" / f"{task.host.name}_ipv4.txt")
                exporter = TextFileExporter(file_path, ipv4_routes)
                exporter.export_to_file()
            else:
                print(f"{Fore.RED}{task.host.name}: No IPv4 routes are defined.")
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")

    def export_ipv6_routes(self, task: Task) -> None:
        """
        Export směrovací IPv6 tabulky jednotlivých síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/ip_routes/{konkrétní host}_ipv6.txt (při spuštění skriptu na GNU/Linux).
        Podporovaní výrobci + typy zařízení: Cisco (L3 switche nebo routery), Juniper (routery).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Raises:
            NornirSubTaskError: Výjimka, která nastane, pokud nastana chyba v nornir úkolu nebo pokud provádíte
                export na nepodporovaných zařízeních.

        Returns:
            None

        """
        command = ""
        if task.host['vendor'] == "cisco" and (
                task.host['dev_type'] == "router" or task.host['dev_type'] == "L3_switch"):
            command = "show ipv6 route"
        elif task.host['vendor'] == "juniper" and task.host['dev_type'] == "router":
            command = "show route"
        else:
            raise NornirSubTaskError("Function was not implemented for particular vendor.", task)
        result = task.run(task=netmiko_send_command, name="Get IP routes", command_string=command)
        if not result.failed:
            ipv6_routes = result[0].result
            if task.host['vendor'] == "juniper":
                parser = NetworkInfoParser()
                ipv6_routes = parser.get_parsed_juniper_routes(result, ipv6_routes=True)
            if ipv6_routes != "":
                file_path = Path(Path.cwd() / 'export' / "ip_routes" / f"{task.host.name}_ipv6.txt")
                exporter = TextFileExporter(file_path, ipv6_routes)
                exporter.export_to_file()
            else:
                print(f"{Fore.RED}{task.host.name}: No IPv6 routes are defined.")
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")

    def export_packet_filter_info(self, task: Task) -> None:
        """
        Export definovaných paketových filtrů (IPv4 i IPv6) u jednotlivých síťových prvků do .txt souborů.
        Export je proveden paralelně. Výsledná cesta je ./export/packet_filter/{konkrétní host}.txt (při spuštění skriptu na GNU/Linux).
        Podporovaní výrobci + typy zařízení: Cisco (L3 switche nebo routery), Juniper (routery).

        Args:
            task (Task): Task objekt, umožňující paralelně volat a seskupovat další nornir úkoly (funkce).

        Raises:
            NornirSubTaskError: Výjimka, která nastane, pokud nastana chyba v nornir úkolu nebo pokud provádíte
                export na nepodporovaných zařízeních.

        Returns:
            None

        """
        command = ""
        if task.host['vendor'] == "cisco" and (
                task.host['dev_type'] == "router" or task.host['dev_type'] == "L3_switch"):
            command = "do show access-lists"
        elif task.host['vendor'] == "juniper" and task.host['dev_type'] == "router":
            command = "show firewall"
        else:
            raise NornirSubTaskError("Function was not implemented for particular vendor.", task)
        result = task.run(task=netmiko_send_config, name="Get packet filters info", config_commands=[command])
        if not result.failed:
            parser = NetworkInfoParser()
            packet_filter_info = parser.get_parsed_packet_filter_data(task.host['vendor'], result)
            if packet_filter_info != "":
                file_path = Path(Path.cwd() / 'export' / "packet_filter" / f"{task.host.name}.txt")
                exporter = TextFileExporter(file_path, packet_filter_info)
                exporter.export_to_file()
            else:
                print(f"{Fore.RED}{task.host.name}: No packet filter is defined.")
        else:
            print(f"{Fore.RED}Export failed for host {task.host.name} more in nornir.log")
